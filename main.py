from dateutil import parser
from datetime import datetime as dt
from dateutil.relativedelta import relativedelta
import xlsxwriter
import ephem
import pytz
import datetime
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import numbers
import tzlocal
import astral

import datetime
from astral import LocationInfo
from astral import Observer
from astral.geocoder import database, lookup
from astral.sun import sun, night
from astral.moon import moonrise, moonset
from astral import Depression

from dateutil.rrule import rrule, DAILY

format_data = "%m/%d/%y %H:%M:%S.%f %Z"

def main():
    ROW_OFFSET = 1

    # Make an observer
    obs = ephem.Observer()

    """
    Disabled lat/lon and date input prompt so you don't have to type them in every time while troubleshooting
    lat = input("Enter your latitude: ")
    lon = input("Enter your longitude: ")
    start_date = parser.parse(input("Enter the start date (YYYY/MM/DD): "))
    end_date = parser.parse(input("Enter the end date (YYYY/MM/DD): "))
    """
    lat = '40.7720'
    lon = '-112.1012'
    start_date = parser.parse('2023/05/01')
    end_date = parser.parse('2023/06/01')

    obs.lat = lat
    obs.lon = lon
    start_utc = ephem.Date(start_date)
    end_utc = ephem.Date(end_date + relativedelta(days=1))

    current_utc = start_utc

    moonset_times = []
    end_twilight_times = []
    begin_twilight_times = []
    moonrise_times = []

    # Note: The Ephem library ALWAYS uses Universal Time (UTC), never the local time zone, which complicates things
    while current_utc < end_utc:
        obs.date = current_utc

        # Calculate the moonrise and moonset times
        obs.horizon = '0'
        moonset_time = obs.next_setting(ephem.Moon(), start=current_utc)
        moonrise_time = obs.previous_rising(ephem.Moon(), start=current_utc)

        # Calculate the astronomical twilight times (when sun is -18 below horizon)
        obs.horizon = '-18'
        begin_twilight = obs.previous_rising(ephem.Sun(), start=current_utc, use_center=True)
        end_twilight = obs.next_setting(ephem.Sun(), start=current_utc, use_center=True)

        # Convert the UTC times to the observer's local timezone
        obs.date = begin_twilight
        begin_twilight_local = ephem.localtime(obs.date).strftime('%Y/%m/%d %I:%M %p')
        obs.date = end_twilight
        end_twilight_local = ephem.localtime(obs.date).strftime('%Y/%m/%d %I:%M %p')
        obs.date = moonrise_time
        moonrise_time_local = ephem.localtime(obs.date).strftime('%Y/%m/%d %I:%M %p')
        obs.date = moonset_time
        moonset_time_local = ephem.localtime(obs.date).strftime('%Y/%m/%d %I:%M %p')

        # Add the local time zone moonset/rise and astronomical twilight begin/end times to the list
        moonset_times.append(moonset_time.datetime())
        end_twilight_times.append(end_twilight.datetime())
        begin_twilight_times.append(begin_twilight.datetime())
        moonrise_times.append(moonrise_time.datetime())

        # Terminal printout to help with troubleshooting
        print('Begin astronomical twilight:', begin_twilight)
        print('End astronomical twilight:', end_twilight)
        print('Moonrise time:', moonrise_time)
        print('Moonset time:', moonset_time)
        print()

        # Increment the current date by one day
        current_utc += 1


    # Create Excel file using xlsxwriter
    workbook = xlsxwriter.Workbook("darkSkyTimes.xlsx")
    worksheet = workbook.add_worksheet('DarkSkyTimes')

    # Format cells
    cell_header_format = workbook.add_format()
    cell_header_format.set_bold()
    cell_header_format.set_text_wrap()
    cell_header_format.set_align('center_across')

    time_format = workbook.add_format()
    time_format.set_text_wrap()
    time_format.set_align('center_across')

    # Create headers using cell location
    worksheet.write('A2', 'Day', cell_header_format)
    worksheet.write('B2', 'Moon Set', cell_header_format)
    worksheet.write('C2', 'Astronomical Twilight END', cell_header_format)
    worksheet.write('D2', 'Astronomical Twilight START', cell_header_format)
    worksheet.write('E2', 'Moon Rise', cell_header_format)
    worksheet.write('F2', 'Duration', cell_header_format)

    rowIndex = 3

    for row, moon_set_sheet in enumerate(moonset_times):
        day_sheet = row + ROW_OFFSET
        moon_set_tz = moon_set_sheet.replace(tzinfo=pytz.timezone('UTC'))
        moon_set = moon_set_tz.astimezone(pytz.timezone('US/Mountain'))
        # Included %Y/%m/%d to help with troubleshooting
        moon_set_str = moon_set.strftime('%Y/%m/%d %H:%M')
        local_moon_set_day = moon_set.astimezone(pytz.timezone('US/Mountain')).day

        """
        On days when the moon does not rise or does not set in that day, then that cell should be blank.
        However, the rise/set time for the next day was being populated in that cell that should be blank.
        This solves that problem, but creates a new problem where the last day or two of the month is
        skipping a day or being displayed on the wrong day.

        Current problem:
        Second to last day of month: incorrect moon rise time, it displays the moon rise time for the next day.
        Last day of month: incorrect moon set time, it displays the time for the next day (which is the first day
        of the next month). Also, no data is being filled in for the last day of the month for the moon rise time,
        because it is being inputted on the second to last day of the month, so that cell is blank when it shouldn't be.
        """
        moon_rise_tz = moonrise_times[row].replace(tzinfo=pytz.timezone('UTC'))
        moon_rise = moon_rise_tz.astimezone(pytz.timezone('US/Mountain'))
        # Included %Y/%m/%d to help with troubleshooting
        moon_rise_str = moon_rise.strftime('%Y/%m/%d %H:%M')
        local_moon_rise_day = moon_rise.astimezone(pytz.timezone('US/Mountain')).day + 1
        moon_rise_month = moon_rise.astimezone(pytz.timezone('US/Mountain')).month
        moon_rise_year = moon_rise.astimezone(pytz.timezone('US/Mountain')).year
        first, last = monthrange(moon_rise_year, moon_rise_month)
        if local_moon_rise_day > last:
            local_moon_rise_day = local_moon_rise_day - last

        # Included %Y/%m/%d to help with troubleshooting
        end_twilight_sheet = end_twilight_times[row].replace(tzinfo=pytz.timezone('UTC')).astimezone(pytz.timezone('US/Mountain')).strftime('%Y/%m/%d %H:%M')
        # Included %Y/%m/%d to help with troubleshooting
        begin_twilight_sheet = begin_twilight_times[row].replace(tzinfo=pytz.timezone('UTC')).astimezone(pytz.timezone('US/Mountain')).strftime('%Y/%m/%d %H:%M')

        worksheet.write('A' + str(rowIndex + 1), day_sheet, cell_header_format)
        if local_moon_set_day != day_sheet:
            worksheet.write('B' + str(rowIndex), moon_set_str, time_format)
        else:
            worksheet.write('B' + str(rowIndex + 1), moon_set_str, time_format)
        worksheet.write('C' + str(rowIndex), end_twilight_sheet, time_format)
        worksheet.write('D' + str(rowIndex), begin_twilight_sheet, time_format)
        if local_moon_rise_day != day_sheet:
            worksheet.write('E' + str(rowIndex - 1), moon_rise_str, time_format)
        else:
            worksheet.write('E' + str(rowIndex), moon_rise_str, time_format)

        rowIndex += 1


    # Set column size
    worksheet.set_column('A:A', 5)
    worksheet.set_column('B:B', 20)
    worksheet.set_column('C:C', 23)
    worksheet.set_column('D:D', 25)
    worksheet.set_column('E:E', 21)
    worksheet.set_column('F:F', 9)

    workbook.close()

    # Load the workbook using openpyxl
    workbook = load_workbook('darkSkyTimes.xlsx')

    # Get the active sheet
    sheet = workbook.active

    # Delete the row of times that is generated before the specified date-range since Ephem uses "previous_rising"
    sheet.delete_rows(3)

    # Save the modified workbook
    workbook.save('darkSkyTimes.xlsx')

class AstroDay:
    def __str__(self):
        return f'The day is: {self.day}, {self.moon_rise = }, {self.moon_set = }, {self.twilight_end = }, {self.twilight_start = }'

    def __init__(self, day):
        self.day = day
        self.moon_rise = None
        self.moon_set = None
        self.twilight_end = None
        self.twilight_start = None

    def populate_astro_data (self, obs):
        # To get the moon set and dusk for today, we need to ask ephem the next setting of the moon and sun
        # starting with midnight today in this timezone.
        midnight_tonight = datetime.datetime.combine(self.day, datetime.datetime.min.time(), tzinfo=tzlocal.get_localzone())
        e_date = ephem.Date(midnight_tonight)
        obs.date = e_date
        obs.horizon = '0'
        moon_set_unaware_dt = obs.next_setting(ephem.Moon()).datetime()
        obs.horizon = '-18'
        dusk_unaware_dt = obs.next_setting(ephem.Sun(), use_center=True).datetime()

        # To get the moon rise and dawn for today, we need to ask ephem for the previous rising
        # of the moon and sun starting with midnight tomorrow in this timezone.
        midnight_tomorrow = datetime.datetime.combine(self.day + datetime.timedelta(days=1), datetime.datetime.min.time(),
                                                     tzinfo=tzlocal.get_localzone())
        e_date = ephem.Date(midnight_tomorrow)
        obs.date = e_date
        obs.horizon = '0'
        moon_rise_unaware_dt = obs.previous_rising(ephem.Moon()).datetime()
        obs.horizon = '-18'
        dawn_unaware_dt = obs.previous_rising(ephem.Sun(), use_center=True).datetime()

        # Calculate the moonrise and moonset times
        # moon_set_unaware_dt = obs.next_setting(ephem.Moon()).datetime()
        # moon_rise_unaware_dt = obs.previous_rising(ephem.Moon()).datetime()

        # Calculate the astronomical twilight times (when sun is -18 below horizon)
        # obs.horizon = '-18'
        # twilight_start_unaware_dt = obs.previous_rising(ephem.Sun(), use_center=True).datetime()
        # twilight_end_unaware_dt = obs.next_setting(ephem.Sun(), use_center=True).datetime()

        self.moon_set = pytz.utc.localize(moon_set_unaware_dt)
        self.moon_rise = pytz.utc.localize(moon_rise_unaware_dt)
        self.dawn = pytz.utc.localize(dawn_unaware_dt)
        self.dusk = pytz.utc.localize(dusk_unaware_dt)


def test():

    print("hello")
    today = datetime.date(year=2023, month=5, day=17)
    d = AstroDay(today)
    print(d)

    # Make an observer
    obs = ephem.Observer()
    lat = '40.7720'
    lon = '-112.1012'
    obs.lat = lat
    obs.lon = lon

    d.populate_astro_data(obs)
    print(d)
    print()
    print('========= Ephem ===========')
    print(f'dawn: {d.dawn.astimezone().strftime(format_data)}')
    print(f'dusk: {d.dusk.astimezone().strftime(format_data)}')
    print(f'moon rise: {d.moon_rise.astimezone().strftime(format_data)}')
    print(f'moon set: {d.moon_set.astimezone().strftime(format_data)}')
    print()

def roundTime(dt=None, roundTo=60):
   """Round a datetime object to any time lapse in seconds
   dt : datetime.datetime object, default now.
   roundTo : Closest number of seconds to round to, default 1 minute.
   Author: Thierry Husson 2012 - Use it as you want but don't blame me.
   """
   if dt == None : dt = datetime.datetime.now()
   seconds = (dt.replace(tzinfo=None) - dt.min).seconds
   rounding = (seconds+roundTo/2) // roundTo * roundTo
   return dt + datetime.timedelta(0,rounding-seconds,-dt.microsecond)

def test_astral():
    city = lookup("Salt Lake City", database())
    # print((
    #     f"Information for {city.name}/{city.region}\n"
    #     f"Timezone: {city.timezone}\n"
    #     f"Latitude: {city.latitude:.02f}; Longitude: {city.longitude:.02f}\n"
    #     f"Elevation: {city.observer.elevation:.02f}" # doesn't have elevation data sadly
    # ))

    lat = 40.7720
    lon = -112.1012
    bird_sanctuary = Observer(lat, lon)

    # don't need to account for elevation https://stackoverflow.com/questions/7662543/results-for-observer-seemingly-not-accounting-for-elevation-effects-in-pyephem

    # print((
    #     f"Latitude: {bird_sanctuary.latitude:.02f}; Longitude: {bird_sanctuary.longitude:.02f}\n"
    #     f"Elevation: {bird_sanctuary.elevation:.02f}"
    # ))

    day = datetime.date(year=2023, month=5, day=30)

    # s = sun(bird_sanctuary, date=day, dawn_dusk_depression=Depression.ASTRONOMICAL)
    s = sun(bird_sanctuary, date=day, dawn_dusk_depression=Depression.ASTRONOMICAL, tzinfo=tzlocal.get_localzone())

    # print('====================')
    # print((
    #     f'Dawn:    {s["dawn"].astimezone().strftime(format_data)}\n'
    #     # f'Sunrise: {s["sunrise"].astimezone().isoformat()}\n'
    #     # f'Noon:    {s["noon"].astimezone().isoformat()}\n'
    #     # f'Sunset:  {s["sunset"].astimezone().isoformat()}\n'
    #     f'Dusk:    {s["dusk"].astimezone().strftime(format_data)}\n'
    # ))

    print(s["dusk"])

    mrise = moonrise(bird_sanctuary, date=day, tzinfo=tzlocal.get_localzone())
    mset = moonset(bird_sanctuary, date=day, tzinfo=tzlocal.get_localzone())
    print('========= Astral ===========')
    print((
        f'Dawn:    {s["dawn"].strftime(format_data)}\n'
        f'Dusk:    {s["dusk"].strftime(format_data)}\n'
        f'Moonrise: {mrise.strftime(format_data)}\n'
        f'Moonset:  {mset.strftime(format_data)}\n'
    ))

    # # this appears to be way off
    # n = night(bird_sanctuary, date=day, tzinfo=tzlocal.get_localzone())
    # print((
    #     f'night start:  {n[0].strftime(format_data)}\n'
    #     f'night end:    {n[1].strftime(format_data)}\n'
    # ))


class AstralDay:
    def __init__(self, day, lat, lon):
        self.day = day
        self.obs = Observer(lat, lon)

        s = sun(self.obs, date=self.day, dawn_dusk_depression=Depression.ASTRONOMICAL, tzinfo=tzlocal.get_localzone())
        self.dawn = s["dawn"]
        self.dusk = s["dusk"]

        try:
            self.moon_rise = moonrise(self.obs, date=self.day, tzinfo=tzlocal.get_localzone())
        except ValueError as e:
            if str(e) != "Moon never rises on this date, at this location":
                raise
            self.moon_rise = None

        try:
            self.moon_set = moonset(self.obs, date=self.day, tzinfo=tzlocal.get_localzone())
        except ValueError as e:
            if str(e) != "Moon never sets on this date, at this location":
                raise
            self.moon_set = None

    def __str__(self):
        dt_fmt = "%I:%M %p %Z"

        if self.moon_rise:
            moon_rise_str = self.moon_rise.strftime(dt_fmt)
        else:
            moon_rise_str = ' -          '

        if self.moon_set:
            moon_set_str = self.moon_set.strftime(dt_fmt)
        else:
            moon_set_str = ' -          '

        return f'| {self.day.strftime("%m/%d/%y")} | {self.dawn.strftime(dt_fmt)} | {self.dusk.strftime(dt_fmt)} | {moon_rise_str} | {moon_set_str} |'


def test_loop_astral():
    lat = 40.7720
    lon = -112.1012

    start = datetime.date(year=2023, month=5, day=1)
    end = datetime.date(year=2023, month=5, day=31)

    print('+----------+--------------+--------------+--------------+--------------+')
    print('| Date     | Dawn         | Dusk         | Moon Rise    | Moon Set     |')
    print('+----------+--------------+--------------+--------------+--------------+')

    for day in rrule(DAILY, dtstart=start, until=end):
        print(AstralDay(day, lat, lon))

def test_17():
    print("====== test missing data on 5/17/23 =========")
    day = datetime.date(year=2023, month=5, day=17)

    slc = lookup("Salt Lake City", database())
    try:
        print(moonset(slc.observer, day, tzinfo=tzlocal.get_localzone()).strftime(format_data))
    except ValueError as e:
        print(f'caught exception: {str(e)}')

if __name__ == "__main__":
#    main()
    test()
    test_astral()
    test_loop_astral()
    test_17()

# TODO
# Throw out or set to none when ephem returns a date time that isn't today.
# Loop through all the dates in our range.
